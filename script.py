import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tqdm import tqdm


def x_y_labels(file_path):
    try:
        book = load_workbook(file_path)
        sheet = book.active
        x_value = sheet.cell(row=1, column=1).value
        y_value = sheet.cell(row=1, column=2).value
    except:
        x_value = "NULL"
        y_value = "NULL"
    return(x_value, y_value)


def main():
    list_of_file = os.listdir('excel_file')
    file_path = 'excel_file/'+list_of_file[0]
    if len(list_of_file) == 0:
        raise Exception("Error! No files found")
    else:
        x_value, y_value = x_y_labels(file_path)
        excel_file = pd.ExcelFile(file_path)
        number_of_sheets = len(excel_file.sheet_names)
        print("File name:", list_of_file[0])
        print("Number of sheets:", number_of_sheets, "\n")
        for sheet_name in tqdm(excel_file.sheet_names):
            sheet_data = pd.read_excel(excel_file, sheet_name)
            df = pd.DataFrame(sheet_data, columns=[x_value,  y_value])
            plt.plot(df[x_value], df[y_value], marker='o')
            plt.title(sheet_name)
            plt.xlabel(x_value)
            plt.ylabel(y_value)
            for i, v in enumerate(df[y_value]):
                plt.text(x=i+0.1, y=v, s=str(v))
            plt.grid(True)
            plt_graph_savepath = "output_graphs/" + sheet_name + ".jpg"
            plt.savefig(plt_graph_savepath, format='JPEG')
            plt.close()


if __name__ == "__main__":
    main()
